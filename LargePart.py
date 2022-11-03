"""
Title: Excel-controlled Large Part Library Component Creator
@Author: Tony Radice
@Date: 10/03/22
@Version:
Purpose: Use specified Excel Workbook to create a large pin-number and
   / or body KiCad Schematic Symbol with attached properties.  Use a formatted
   .xlsx file with name PDef_{component name} to output file
   {component name}.kicad_sym suitable for importing into KiCad Symbol Editor.

"""

"""
0.1 - Generated a Correct Single Body Part (saved)
0.2 - Target: Multi Body Part (saved)
0.21 - Add Body Count, Pin Count and Duplicate Pin Name Check
0.3 - Rewrite Write_Top_Block and Get_Top_Args to allow for additional
      symbol level properties (See input spreadsheet).  Rewrite Parse_
      command line
"""

# External Resources Definition -
import os
import sys
import time
import openpyxl
from LP_lib import *

# Admin 02 - Static Variable Declarations
dcode = time.strftime("%Y%m%d")
DPL = []  # Duplicate Pin List - len is number of pins created

Arguments = ParseCommandLine(sys.argv)
DefFile = Arguments[0] + '.xlsx'                # Input Description File
Debug01 = Arguments[1]                          # Debug Flag

if Debug01: print(dcode)

# Check for Master Excel Sheet
# print('LP-D-Opening', DefFile, 'Spreadsheet.')
if not os.path.exists(DefFile):
    exit("LargePart-E-Part Definition File not found.")

# Create an Output File Name (.kicad_sym) - Split off PreFix
tofn = Arguments[0].rsplit("_")
OutputFileName = (tofn[1] + '.kicad_sym')
# print('LP-D-Output File Name:',OutputFileName)
OFile = open(OutputFileName, 'wt')

# Open Part Definition Workbook
ewb = openpyxl.load_workbook(DefFile, data_only=True)
LibsList = ewb.sheetnames
# if Debug01: print('LP-D-Sheets List:', LibsList)
TopIndex = LibsList.index('Top')
# if Debug01: print('LP-D-Top Index: ',TopIndex)

# Generate the Top section of the Output
TopWorksheet = ewb["Top"]
# ToDo Throw an Error if there is no sheet named Top

# ToDo Insert an error check routine for Top Page
if CheckTopSheet(TopWorksheet):
    print('LP-E- Top Sheet Error.')

WriteTopBlock(TopWorksheet, OFile)              # Write Common Symbol properties
PartName = GetPartName(TopWorksheet)            # Need Part name for bodies

# ToDo - Possible Error if Top sheet is not first in list!
LibsList.pop(0)  # Remove the Top Sheet from List

# if Debug01: print('LP-D-Sheets Extracted List:', LibsList)

BodyNumber = 0
for Sheet in LibsList:
    CurrentWorkSheet = ewb[Sheet]
    #    if Debug01: print('LP-D-Tab:', CurrentWorkSheet)

    if CurrentWorkSheet.cell(row=1, column=1).value == 'DNI':
        print('LP-I- Skip worksheet:', Sheet)

    else:
        # ToDo Insert an error check routine for Body Pages
#        if CheckDataSheet(CurrentWorkSheet):
#            print('LP-E- Data Sheet Error.' + Sheet)

        BodyNumber += 1
        MaxH = MaxHeight(CurrentWorkSheet)
        #    if Debug01: print('LP-D-Tab MaxHeight:', MaxH)

        #   Need to append a Body Number to the Part Name to pass
        LocalPartName = PartName + '_' + str(BodyNumber)
        #    if Debug01: print('LP-D-PartName:', LocalPartName)
        CreateSymbolBodyHeader(LocalPartName, CurrentWorkSheet, MaxH, OFile)

        #   Place all Pins
        DPLL = PlacePinsInBody(CurrentWorkSheet, MaxH, OFile)
        DPL.extend(DPLL)  # Probably cumbersome way of doing this...
        SymbolClose(OFile)

if not CheckForDupPins(DPL):                    # A Duplicate returns True
    print('LP-I- Duplicates Check passed.')

# End of Program
print('LP-I- ' + str(BodyNumber) + ' Body(s) created.')
print('LP-I- ' + str(len(DPL)) + ' Pins created.')

TempClose(OFile)
OFile.close()
